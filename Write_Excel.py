import openpyxl as op

class Excel:
    stocks = []
    watchlist = []
    filename = ""
    sheetname = ""
    def __init__(self, stocks, watchlist, filename, sheetname):
        self.stocks = stocks
        self.watchlist = watchlist
        self.filename = filename
        self.sheetname = sheetname


    def write(self):

        workfile = op.load_workbook(self.filename)
        sheet = workfile[self.sheetname]

        row = 1;
        columns = []
        for i in range(1, sheet.max_column+1):
            columns.append(sheet.cell(row, i).value)


        watchlist_column = columns.index('Watchlist_Name') + 1
        stock_column = columns.index('Stock_Name') + 1
        # nse_column = columns.index('NSE_Name') + 1
        # price_column = columns.index('CMP') + 1

        row+=1
        for i in range(0, len(self.stocks)):
            watchlist_cell = sheet.cell(row, watchlist_column)
            watchlist_cell.value = self.watchlist[i]
            for name in self.stocks[i]:
                stock_cell = sheet.cell(row, stock_column)
                stock_cell.value = name
                row+=1

        workfile.save('Stocks.xlsx')

        
        



